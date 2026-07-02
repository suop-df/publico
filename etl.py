"""
ETL: Oracle (BO/GDF) -> JSON.GZ -> GitHub Releases
Extrai dados via oracledb, comprime com gzip e salva em data/gz/
para upload no Release 'dados-latest' do GitHub.
Os dashboards consomem os arquivos diretamente do Release.
"""

import os
import json
import gzip
import logging
from datetime import datetime, date, timezone
from decimal import Decimal
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()

DB_USER     = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")
DB_DSN      = os.getenv("DB_DSN", "10.69.1.118:1521/oraprd06")
CLIENT_PATH = os.getenv("ORACLE_CLIENT_PATH", "").strip()
FCDF_PATH   = os.getenv("FCDF_PATH", "").strip()
DB_MIN  = int(os.getenv("DB_MIN_CONNECTIONS", 1))
DB_MAX  = int(os.getenv("DB_MAX_CONNECTIONS", 5))
DB_INC  = int(os.getenv("DB_INCREMENT_CONNECTIONS", 1))

SUPABASE_URL = os.getenv("SUPABASE_URL", "").strip()
SUPABASE_KEY = os.getenv("SUPABASE_KEY", "").strip()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)

BASE_DIR    = Path(__file__).parent
OUTPUT_DIR  = BASE_DIR / "data"
GZ_DIR      = BASE_DIR / "data" / "gz"
QUERIES_DIR = BASE_DIR / "data" / "queries"
OUTPUT_DIR.mkdir(exist_ok=True)
GZ_DIR.mkdir(exist_ok=True)

SCHEMA_ANO = f"mil{datetime.now().year}"

QUERIES = [
    {
        "file": "receita.json",
        "sql_file": "RECEITA.sql",
    },
    {
        "file": "despesa.json",
        "sql_file": "DESPESA.sql",
    },
    {
        "file": "rcl.json",
        "sql_file": "receita_RCL.sql",
        "transform": "rcl",
    },
    {
        "file": "restos_a_pagar.json",
        "sql_file": "restos_a_pagar.sql",
        "transform": "restos_a_pagar",
    },
    {
        "file": "resultado_primario_nominal.json",
        "sql_file": "resultado_primario_nominal.sql",
        "transform": "resultado_primario_nominal",
    },
    {
        "file": "poupanca_corrente.json",
        "sql_file": "poupanca_corrente.sql",
        "transform": "poupanca_corrente",
    },
    {
        "file": "minimo_saude.json",
        "sql_file": "minimo_saude.sql",
        "transform": "minimo_saude",
    },
    {
        "file": "minimo_educacao.json",
        "sql_file": "minimo_educacao.sql",
        "transform": "minimo_educacao",
    },
]


def serialize(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def fetch(cursor, query):
    cursor.execute(query)
    columns = [col[0].lower() for col in cursor.description]
    return [
        {col: serialize(val) for col, val in zip(columns, row)}
        for row in cursor.fetchall()
    ]


def save_json(filename, data):
    payload = {
        "atualizado_em": datetime.now(timezone.utc).isoformat(),
        "total": len(data),
        "dados": data,
    }
    path = OUTPUT_DIR / filename
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    log.info(f"  {filename} -- {len(data)} registros salvos")


def save_json_gz(filename, data):
    payload = {
        "atualizado_em": datetime.now(timezone.utc).isoformat(),
        "total": len(data),
        "dados": data,
    }
    content = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    gz_filename = filename + ".gz"
    path = GZ_DIR / gz_filename
    with gzip.open(path, "wb", compresslevel=9) as f:
        f.write(content)
    size_kb = path.stat().st_size / 1024
    log.info(f"  {gz_filename} -- {len(data)} registros, {size_kb:.1f} KB comprimido")


def read_sql(filename):
    path = QUERIES_DIR / filename
    if not path.exists():
        raise FileNotFoundError(f"Arquivo SQL nao encontrado: {path}")
    sql = path.read_text(encoding="utf-8")
    lines = [line for line in sql.splitlines() if not line.strip().startswith("--")]
    return (
        "\n".join(lines)
        .replace("{SCHEMA_ANO}", SCHEMA_ANO)
        .strip()
    )


def resolve_query(item):
    return read_sql(item["sql_file"])


# RCL Aggregation
# Regras: tools/regra_rcl.txt (ContDF/SEEC)
# FCDF: data/UFIS-FCDFDespesadePessoal.xlsx (UFIS/SIAFE)

FCDF_CLASS6_CONTAS = {622130300, 622130400, 622130500, 622130600, 622130700}

# RREO Anexo 06 — Juros Ativos (XXXVI) e Passivos (XXXVII)
_XXXVI_CC5 = {
    "44121","44123","44124","44125","44131","44133","44134","44135","44141",
    "44211","44213","44214","44215","44221","44261","44263","44264","44265",
    "44511","44521","44611","44613","44614","44615","44621","44623","44624","44625",
}
_XXXVI_CC7 = {
    "4411199","4431101","4431199","4431301","4431401","4431501",
    "4432101","4433101","4433199","4433301","4433401","4433501",
    "4434101","4435101","4435301","4435401","4435501",
}
_XXXVI_FULL = {"443910170", "443930170", "443930171"}

_XXXVII_CC5 = {
    "34111","34113","34114","34115","34121","34131","34133","34134","34135","34141",
    "34181","34183","34184","34185","34191","34211","34213","34214","34215","34221",
    "34261","34263","34264","34265","34511","34521","34611","34613","34614","34615",
    "34911","34913","34914","34915",
}
_XXXVII_CC7 = {
    "3425202","3431101","3431301","3431401","3431501","3432101","3433101",
    "3433301","3433401","3433501","3434101","3435101","3435301","3435401","3435501",
}
_XXXVII_FULL = {"343910170", "343930170", "343930171"}


def _rcl_class_orc(c, cofonte, cofontefederal):
    if 11125000 <= c <= 11125099: return "iptu"
    if 11130000 <= c <= 11139999: return "ir"
    if 11125100 <= c <= 11125199: return "ipva"
    if 11125200 <= c <= 11125299: return "itcd"
    if 11125300 <= c <= 11125399: return "itbi"
    if (11145010 <= c <= 11145099) or (11145200 <= c <= 11145299): return "icms"
    if 11145100 <= c <= 11145199: return "iss"
    if 11190000 <= c <= 11199999: return "outros_impostos"
    if 11200000 <= c <= 11299999: return "taxas"
    if 12000000 <= c <= 12999999: return "contribuicoes"
    if 13200000 <= c <= 13299999: return "rend_aplic"
    if (13100000 <= c <= 13199999) or (13300000 <= c <= 13999999): return "outras_patrimoniais"
    if 14000000 <= c <= 14999999: return "agropecuaria"
    if 15000000 <= c <= 15999999: return "industrial"
    if 16000000 <= c <= 16999999: return "servicos"
    if 17115000 <= c <= 17115099: return "fpe"
    if 17115100 <= c <= 17115199: return "fpm"
    if 17115200 <= c <= 17115299: return "itr_trans"
    if 17115300 <= c <= 17115399: return "lc61"
    if (17515000 <= c <= 17515099) or (17155200 <= c <= 17155299): return "fundeb_trans"
    if (17115400 <= c <= 17155199) or (17155300 <= c <= 17514999) or (17515100 <= c <= 17999999):
        return "outras_transf"
    if 19000000 <= c <= 19999999: return "outras_correntes"
    return None


def _rcl_deducao(c):
    if 12150000 <= c <= 12159999: return "contrib_servidor"
    if 19990300 <= c <= 19990399: return "comp_financeira"
    if 13210400 <= c <= 13210499: return "rend_prev"
    if 17515000 <= c <= 17515099: return "ded_fundeb"
    return None


def _is_xxxvi(cc):
    return cc[:5] in _XXXVI_CC5 or cc[:7] in _XXXVI_CC7 or cc in _XXXVI_FULL


def _is_xxxvii(cc):
    return cc[:5] in _XXXVII_CC5 or cc[:7] in _XXXVII_CC7 or cc in _XXXVII_FULL


def _rcl_emenda(cofonte, cofontefederal):
    if (732000000 <= cofonte <= 732999999 or
            738000000 <= cofonte <= 738999999 or
            706000000 <= cofonte <= 706999999):
        return "emendas_ind"
    if (733000000 <= cofonte <= 733999999 or
            739000000 <= cofonte <= 739999999):
        return "emendas_bancada"
    if cofontefederal == 1604:
        return "agentes_com"
    return None


def load_fcdf_data(base_dir):
    if FCDF_PATH:
        path = Path(FCDF_PATH)
    else:
        path = base_dir / "data" / "UFIS-FCDFDespesadePessoal.xlsx"
    if not path.exists():
        log.warning(f"Planilha FCDF nao encontrada: {path}")
        return {"realizados": {}, "previsao": {}}
    try:
        from openpyxl import load_workbook
    except ImportError:
        log.error("openpyxl nao instalado. Execute: pip install openpyxl")
        return {"realizados": {}, "previsao": {}}
    from collections import defaultdict
    try:
        wb = load_workbook(path, read_only=True)
    except Exception as e:
        log.warning(f"Planilha FCDF nao pode ser lida ({path}): {e}")
        return {"realizados": {}, "previsao": {}}
    ws = wb.active
    realizados = defaultdict(lambda: {"total": 0.0, "pessoal": 0.0})
    previsao   = defaultdict(lambda: {"total": 0.0, "pessoal": 0.0})
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] is None:
            continue
        try:
            conta  = int(row[2])
            grupo  = int(row[4] or 0)
            elem   = str(row[5]).strip() if row[5] is not None else ""
            subit  = int(row[6]) if row[6] is not None else 0
            mes    = int(row[7] or 0)
            ano    = int(row[8] or 0)
            vadeb  = float(row[9] or 0)
            vacred = float(row[10] or 0)
        except (ValueError, TypeError):
            continue
        if not mes or not ano:
            continue
        is_pessoal = (grupo == 1) or (grupo == 3 and elem == "85" and subit == 1)
        sc = str(conta)
        if sc.startswith("6") and conta in FCDF_CLASS6_CONTAS:
            saldo = vacred - vadeb
            realizados[(mes, ano)]["total"]   += saldo
            if is_pessoal:
                realizados[(mes, ano)]["pessoal"] += saldo
        elif sc.startswith("5"):
            saldo = vadeb - vacred
            previsao[ano]["total"]   += saldo
            if is_pessoal:
                previsao[ano]["pessoal"] += saldo
    wb.close()
    log.info(f"  FCDF: {len(realizados)} meses realizados, {len(previsao)} anos previsao")
    return {"realizados": dict(realizados), "previsao": dict(previsao)}


def build_rcl_data(rows):
    from collections import defaultdict
    MESES_PT = ["jan","fev","mar","abr","mai","jun",
                "jul","ago","set","out","nov","dez"]
    ano_atual = datetime.now().year
    PREV_EXCL = {521120101, 521220101, 521220201}
    prev_receita  = defaultdict(float)
    prev_deducoes = defaultdict(float)
    prev_emendas  = defaultdict(float)
    real_receita  = defaultdict(lambda: defaultdict(float))
    real_deducoes = defaultdict(lambda: defaultdict(float))
    real_emendas  = defaultdict(lambda: defaultdict(float))
    meses_oracle  = set()

    for r in rows:
        saldo = float(r.get("saldo") or 0)
        if not saldo:
            continue
        cc_raw = r.get("cocontacontabil")
        cc = int(str(cc_raw).strip()) if cc_raw is not None else 0
        class_orc_raw = r.get("class_orc") or ""
        try:
            c_int = int(str(class_orc_raw).strip())
        except ValueError:
            continue
        if not str(class_orc_raw).strip().startswith("1"):
            continue
        mes = int(r.get("inmes") or 0)
        ano = int(r.get("coexercicio") or 0)
        if not mes or not ano:
            continue
        cofonte_raw   = r.get("cofonte")
        cofederal_raw = r.get("cofontefederal")
        try:
            cofonte   = int(str(cofonte_raw).strip()) if cofonte_raw else 0
            cofederal = int(str(cofederal_raw).strip()) if cofederal_raw else 0
        except (ValueError, TypeError):
            cofonte = cofederal = 0

        is_prev = (521100000 <= cc <= 521299999)
        is_real = (621200000 <= cc <= 621399999 and cc != 621310100)

        if is_prev:
            is_excl = (cc in PREV_EXCL)
            if not is_excl:
                key = _rcl_class_orc(c_int, cofonte, cofederal)
                if key and ano == ano_atual:
                    prev_receita[key] += saldo
                ded = _rcl_deducao(c_int)
                # ded_fundeb na previsao vem APENAS das contas excluidas (PREV_EXCL)
                if ded and ded != "ded_fundeb" and ano == ano_atual:
                    prev_deducoes[ded] += saldo
                em = _rcl_emenda(cofonte, cofederal)
                if em and ano == ano_atual:
                    prev_emendas[em] += saldo
            else:
                if 17515000 <= c_int <= 17515099 and ano == ano_atual:
                    prev_deducoes["ded_fundeb"] += saldo
        elif is_real:
            meses_oracle.add((mes, ano))
            key = _rcl_class_orc(c_int, cofonte, cofederal)
            if key:
                real_receita[(mes, ano)][key] += saldo
            ded = _rcl_deducao(c_int)
            if ded:
                real_deducoes[(mes, ano)][ded] += saldo
            em = _rcl_emenda(cofonte, cofederal)
            if em:
                real_emendas[(mes, ano)][em] += saldo

    if not meses_oracle:
        log.warning("RCL: nenhum dado realizado encontrado no SQL.")
        return {}

    # ref_mes/ref_ano: ultimo mes fechado conforme {SCHEMA_ANO}.mesfechado.
    # max_mes_fechado vem como coluna escalar no SQL (mesmo valor em todas as linhas).
    # Fallback: se NULL (mesfechado vazio no ano corrente), usa max dos realizados Oracle.
    mmf_raw = next((r.get("max_mes_fechado") for r in rows if r.get("max_mes_fechado") is not None), None)
    if mmf_raw is not None and 1 <= int(mmf_raw) <= 12:
        ref_mes = int(mmf_raw)
        ref_ano = ano_atual
        log.info(f"  RCL: ultimo mes fechado (mesfechado) = {ref_mes:02d}/{ref_ano}")
    else:
        # Fallback: max dos realizados Oracle
        max_mes, max_ano = max(meses_oracle)
        ref_mes, ref_ano = max_mes, max_ano
        log.warning(f"  RCL: mesfechado vazio/invalido, fallback Oracle max = {ref_mes:02d}/{ref_ano}")

    ultimos12 = []
    m, a = ref_mes, ref_ano
    for _ in range(12):
        ultimos12.insert(0, (m, a))
        m -= 1
        if m == 0:
            m = 12
            a -= 1

    # Janela completa: Jan/ano_anterior até ref_mes/ref_ano (suporte a bimestres)
    ano_anterior = ref_ano - 1
    todas_colunas_ma = ([(mm, ano_anterior) for mm in range(1, 13)] +
                        [(mm, ref_ano) for mm in range(1, ref_mes + 1)])

    janela_padrao = [f"{m},{a}" for m, a in ultimos12]
    colunas  = [f"{m},{a}" for m, a in todas_colunas_ma]
    rotulos  = [f"{MESES_PT[m-1]}/{str(a)[2:]}" for m, a in todas_colunas_ma]
    log.info(f"  RCL: colunas {rotulos[0]} -> {rotulos[-1]} ({len(colunas)} meses)")

    def monta_linha(src, key):
        linha = {}
        for m, a in todas_colunas_ma:
            col = f"{m},{a}"
            linha[col] = src.get((m, a), {}).get(key, 0.0)
        linha["_total"] = sum(linha.get(c, 0.0) for c in janela_padrao)
        return linha

    def soma_linhas(linhas_dict, keys):
        linha = {}
        for col in colunas:
            linha[col] = sum(linhas_dict.get(k, {}).get(col, 0.0) for k in keys)
        linha["_total"] = sum(linha.get(c, 0.0) for c in janela_padrao)
        return linha

    KEYS_ATOMICAS = [
        "iptu","ir","ipva","itcd","itbi","icms","iss","outros_impostos",
        "taxas","itr","contribuicoes","rend_aplic","outras_patrimoniais",
        "agropecuaria","industrial","servicos",
        "fpe","fpm","itr_trans","lc61","fundeb_trans","outras_transf",
        "outras_correntes",
    ]
    IMPOSTOS_KEYS = {"iptu","ir","ipva","itcd","itbi","icms","iss","outros_impostos","taxas","itr"}
    PATRIM_KEYS   = {"rend_aplic","outras_patrimoniais"}
    TRANSF_KEYS   = {"fpe","fpm","itr_trans","lc61","fundeb_trans","outras_transf"}
    CORR_KEYS     = IMPOSTOS_KEYS | PATRIM_KEYS | TRANSF_KEYS | {
                        "contribuicoes","agropecuaria","industrial","servicos","outras_correntes"}

    linhas = {}
    for key in KEYS_ATOMICAS:
        linhas[key] = monta_linha(real_receita, key)
    linhas["itr"] = {col: 0.0 for col in colunas}
    linhas["itr"]["_total"] = 0.0

    linhas["impostos"]           = soma_linhas(linhas, IMPOSTOS_KEYS)
    linhas["patrimonial"]        = soma_linhas(linhas, PATRIM_KEYS)
    linhas["transferencias"]     = soma_linhas(linhas, TRANSF_KEYS)
    linhas["receitas_correntes"] = soma_linhas(linhas, CORR_KEYS)

    for dk in ("contrib_servidor","comp_financeira","rend_prev","ded_fundeb"):
        linhas[dk] = monta_linha(real_deducoes, dk)
    linhas["deducoes"] = soma_linhas(linhas,
        {"contrib_servidor","comp_financeira","rend_prev","ded_fundeb"})

    fcdf      = load_fcdf_data(BASE_DIR)
    real_fcdf = fcdf.get("realizados", {})
    prev_fcdf = fcdf.get("previsao", {})

    def monta_fcdf(campo):
        linha = {}
        for m, a in todas_colunas_ma:
            col = f"{m},{a}"
            linha[col] = real_fcdf.get((m, a), {}).get(campo, 0.0)
        linha["_total"] = sum(linha.get(c, 0.0) for c in janela_padrao)
        return linha

    linhas["fcdf_total"]   = monta_fcdf("total")
    linhas["fcdf_pessoal"] = monta_fcdf("pessoal")
    linhas["fcdf"] = {}
    for col in colunas:
        linhas["fcdf"][col] = (linhas["fcdf_total"].get(col, 0)
                               - linhas["fcdf_pessoal"].get(col, 0))
    linhas["fcdf"]["_total"] = (linhas["fcdf_total"]["_total"]
                                - linhas["fcdf_pessoal"]["_total"])

    linhas["rcl"] = {}
    for col in colunas:
        linhas["rcl"][col] = (linhas["receitas_correntes"].get(col, 0)
                              - linhas["deducoes"].get(col, 0)
                              + linhas["fcdf"].get(col, 0))
    linhas["rcl"]["_total"] = (linhas["receitas_correntes"]["_total"]
                               - linhas["deducoes"]["_total"]
                               + linhas["fcdf"]["_total"])

    for em in ("emendas_ind","emendas_bancada","agentes_com"):
        linhas[em] = monta_linha(real_emendas, em)
    linhas["outras_ded"] = {col: 0.0 for col in colunas}
    linhas["outras_ded"]["_total"] = 0.0

    linhas["rcl_endiv"] = {}
    for col in colunas:
        linhas["rcl_endiv"][col] = (linhas["rcl"].get(col, 0)
                                    - linhas["emendas_ind"].get(col, 0))
    linhas["rcl_endiv"]["_total"] = (linhas["rcl"]["_total"]
                                     - linhas["emendas_ind"]["_total"])

    linhas["rcl_pessoal"] = {}
    for col in colunas:
        linhas["rcl_pessoal"][col] = (
            linhas["rcl_endiv"].get(col, 0)
            - linhas["emendas_bancada"].get(col, 0)
            - linhas["agentes_com"].get(col, 0)
            - linhas["outras_ded"].get(col, 0))
    linhas["rcl_pessoal"]["_total"] = (
        linhas["rcl_endiv"]["_total"]
        - linhas["emendas_bancada"]["_total"]
        - linhas["agentes_com"]["_total"]
        - linhas["outras_ded"]["_total"])

    def pv(key): return prev_receita.get(key, 0.0)
    def pv_g(keys): return sum(prev_receita.get(k, 0.0) for k in keys)

    previsao = {}
    for k in KEYS_ATOMICAS:
        previsao[k] = pv(k)
    previsao["itr"]                = 0.0
    previsao["impostos"]           = pv_g(IMPOSTOS_KEYS)
    previsao["patrimonial"]        = pv_g(PATRIM_KEYS)
    previsao["transferencias"]     = pv_g(TRANSF_KEYS)
    previsao["receitas_correntes"] = pv_g(CORR_KEYS)

    for dk in ("contrib_servidor","comp_financeira","rend_prev","ded_fundeb"):
        previsao[dk] = prev_deducoes.get(dk, 0.0)
    previsao["deducoes"] = sum(prev_deducoes.values())

    pf = prev_fcdf.get(ano_atual, {})
    previsao["fcdf_total"]   = pf.get("total", 0.0)
    previsao["fcdf_pessoal"] = pf.get("pessoal", 0.0)
    previsao["fcdf"]         = previsao["fcdf_total"] - previsao["fcdf_pessoal"]

    previsao["rcl"] = (previsao["receitas_correntes"]
                       - previsao["deducoes"]
                       + previsao["fcdf"])

    for em in ("emendas_ind","emendas_bancada","agentes_com"):
        previsao[em] = prev_emendas.get(em, 0.0)
    previsao["outras_ded"] = 0.0

    previsao["rcl_endiv"]   = previsao["rcl"] - previsao["emendas_ind"]
    previsao["rcl_pessoal"] = (previsao["rcl_endiv"]
                               - previsao["emendas_bancada"]
                               - previsao["agentes_com"]
                               - previsao["outras_ded"])

    return {
        "atualizado_em": datetime.now(timezone.utc).isoformat(),
        "ano": ano_atual,
        "ref_mes": ref_mes,
        "ref_ano": ref_ano,
        "colunas": colunas,
        "rotulos": rotulos,
        "janela_padrao": janela_padrao,
        "linhas": linhas,
        "previsao": previsao,
    }


def build_restos_a_pagar_data(rows):
    """
    Agrega as linhas brutas do SQL por (ano, coug, cocontacontabil, cat, gnd, inmes)
    e retorna lista de registros prontos para o dashboard.
    Saldo = VACREDITO - VADEBITO (contas classe 6 de execução de RAP).
    """
    agg  = {}   # chave -> saldo acumulado
    meta = {}   # chave -> campos descritivos

    for r in rows:
        try:
            cc    = int(str(r.get("cocontacontabil") or 0).strip())
            saldo = float(r.get("vacredito") or 0) - float(r.get("vadebito") or 0)
            inmes = int(r.get("inmes") or 0)
            ano   = int(r.get("ano") or 0)
        except (ValueError, TypeError):
            continue
        if not ano:
            continue

        coug = str(r.get("coug") or "").strip()
        cat  = str(r.get("cocategoriaeconomica") or "").strip()
        gnd  = str(r.get("cognd") or "").strip()
        key  = (ano, coug, cc, cat, gnd, inmes)

        agg[key]  = agg.get(key, 0.0) + saldo
        meta[key] = {
            "noug":  str(r.get("noug")  or "").strip(),
            "nocat": str(r.get("nocategoriaeconomica") or "").strip(),
            "nognd": str(r.get("nognd") or "").strip(),
        }

    registros = []
    for (ano, coug, cc, cat, gnd, inmes), saldo in agg.items():
        m = meta[(ano, coug, cc, cat, gnd, inmes)]
        registros.append({
            "ano":             ano,
            "coug":            coug,
            "noug":            m["noug"],
            "cocontacontabil": cc,
            "cat":             cat,
            "nocat":           m["nocat"],
            "gnd":             gnd,
            "nognd":           m["nognd"],
            "saldo":           round(saldo, 2),
            "inmes":           inmes,
        })

    log.info(f"  Restos a Pagar: {len(registros)} registros agregados")
    return registros


def _supabase_upsert(table, payload, on_conflict, batch_size=1000):
    """Função genérica de upsert no Supabase via REST API."""
    import urllib.request, urllib.error
    url = f"{SUPABASE_URL}/rest/v1/{table}?on_conflict={on_conflict}"
    headers = {
        "apikey":        SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type":  "application/json",
        "Prefer":        "resolution=merge-duplicates",
    }
    total = 0
    for i in range(0, len(payload), batch_size):
        body = json.dumps(payload[i:i+batch_size], ensure_ascii=False).encode("utf-8")
        req  = urllib.request.Request(url, data=body, headers=headers, method="POST")
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                resp.read()
            total += len(payload[i:i+batch_size])
        except urllib.error.HTTPError as e:
            body_err = e.read().decode("utf-8", errors="replace")
            log.error(f"  Supabase [{table}] lote {i}: HTTP {e.code} - {body_err}")
            raise
    return total


def upsert_restos_a_pagar_supabase(registros):
    """Envia restos a pagar para o Supabase."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        log.warning("  Supabase: nao configurado. Pulando upsert restos_a_pagar.")
        return
    try:
        atualizado_em = datetime.now(timezone.utc).isoformat()
        payload = [{**r, "atualizado_em": atualizado_em} for r in registros]
        total = _supabase_upsert("restos_a_pagar", payload,
                                  "ano,coug,cocontacontabil,cat,gnd,inmes")
        log.info(f"  Supabase: {total} registros enviados para restos_a_pagar.")
    except Exception as e:
        log.error(f"  Supabase restos_a_pagar falhou: {type(e).__name__}: {e}")


def upsert_receita_supabase(data):
    """Envia receita orçamentária para o Supabase."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        log.warning("  Supabase: nao configurado. Pulando upsert receita.")
        return
    try:
        atualizado_em = datetime.now(timezone.utc).isoformat()
        payload = []
        for r in data:
            row = {k: (v.strip() if isinstance(v, str) else v) for k, v in r.items()}
            row["atualizado_em"] = atualizado_em
            payload.append(row)
        total = _supabase_upsert("receita", payload,
                                  "coexercicio,inmes,coug,cocontacontabil,cocontacorrente")
        log.info(f"  Supabase: {total} registros enviados para receita.")
    except Exception as e:
        log.error(f"  Supabase receita falhou: {type(e).__name__}: {e}")


def upsert_despesa_supabase(data):
    """Envia despesa orçamentária para o Supabase (com deduplicação por chave única)."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        log.warning("  Supabase: nao configurado. Pulando upsert despesa.")
        return
    try:
        atualizado_em = datetime.now(timezone.utc).isoformat()
        # Deduplicar pela chave única, somando valores financeiros
        agg = {}
        meta = {}
        for r in data:
            row = {k: (v.strip() if isinstance(v, str) else v) for k, v in r.items()}
            row["cofonte"] = row.get("cofonte") or ""
            key = (row.get("coexercicio"), row.get("inmes"), row.get("coug"),
                   row.get("cocontacontabil"), row.get("despesa"), row.get("cofonte"))
            if key not in agg:
                agg[key]  = {"vadebito": 0.0, "vacredito": 0.0, "saldo": 0.0}
                meta[key] = row
            agg[key]["vadebito"]  += float(row.get("vadebito")  or 0)
            agg[key]["vacredito"] += float(row.get("vacredito") or 0)
            agg[key]["saldo"]     += float(row.get("saldo")     or 0)

        payload = []
        for key, vals in agg.items():
            row = {**meta[key], **vals, "atualizado_em": atualizado_em}
            payload.append(row)

        log.info(f"  Despesa: {len(data)} linhas Oracle -> {len(payload)} registros únicos.")
        total = _supabase_upsert("despesa", payload,
                                  "coexercicio,inmes,coug,cocontacontabil,despesa,cofonte",
                                  batch_size=1000)
        log.info(f"  Supabase: {total} registros enviados para despesa.")
    except Exception as e:
        log.error(f"  Supabase despesa falhou: {type(e).__name__}: {e}")


def upsert_resultado_primario_nominal_supabase(D_obj):
    """Envia resultado_primario_nominal para o Supabase como JSONB (1 linha por ano)."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        log.warning("  Supabase: nao configurado. Pulando upsert resultado_primario_nominal.")
        return
    try:
        ano = datetime.now().year
        payload = [{
            "ano":           ano,
            "dados":         D_obj,
            "atualizado_em": datetime.now(timezone.utc).isoformat(),
        }]
        total = _supabase_upsert("resultado_primario_nominal", payload, "ano")
        log.info(f"  Supabase: resultado_primario_nominal {ano} enviada ({total} linha).")
    except Exception as e:
        log.error(f"  Supabase resultado_primario_nominal falhou: {type(e).__name__}: {e}")


def upsert_rcl_supabase(D_obj):
    """Envia RCL para o Supabase como JSONB (1 linha por ano)."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        log.warning("  Supabase: nao configurado. Pulando upsert rcl.")
        return
    try:
        ano = D_obj.get("ano") or datetime.now().year
        payload = [{
            "ano":          ano,
            "dados":        D_obj,
            "atualizado_em": datetime.now(timezone.utc).isoformat(),
        }]
        total = _supabase_upsert("rcl", payload, "ano")
        log.info(f"  Supabase: RCL {ano} enviada ({total} linha).")
    except Exception as e:
        log.error(f"  Supabase rcl falhou: {type(e).__name__}: {e}")


def save_restos_a_pagar_gz(registros):
    payload = {
        "atualizado_em": datetime.now(timezone.utc).isoformat(),
        "registros": registros,
    }
    content = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    gz_path = GZ_DIR / "restos_a_pagar.json.gz"
    with gzip.open(gz_path, "wb", compresslevel=9) as f:
        f.write(content)
    size_kb = gz_path.stat().st_size / 1024
    log.info(f"  restos_a_pagar.json.gz -- {len(registros)} registros, {size_kb:.1f} KB")


def save_resultado_primario_nominal_gz(D_obj):
    payload = {"atualizado_em": datetime.now(timezone.utc).isoformat()}
    payload.update(D_obj)
    content = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    gz_path = GZ_DIR / "resultado_primario_nominal.json.gz"
    with gzip.open(gz_path, "wb", compresslevel=9) as f:
        f.write(content)
    size_kb = gz_path.stat().st_size / 1024
    log.info(f"  resultado_primario_nominal.json.gz -- {size_kb:.1f} KB")


def save_rcl_gz(D_obj):
    content = json.dumps(D_obj, ensure_ascii=False, separators=(",",":")).encode("utf-8")
    gz_path = GZ_DIR / "rcl.json.gz"
    with gzip.open(gz_path, "wb", compresslevel=9) as f:
        f.write(content)
    size_kb = gz_path.stat().st_size / 1024
    log.info(f"  rcl.json.gz -- {len(D_obj.get('colunas', []))} meses, {size_kb:.1f} KB")


def build_poupanca_corrente_data(rows):
    """
    Computa Poupança Corrente (Art. 167-A, CF) com janela móvel de 12 meses.
    Formula: (Desp. Liquidadas 12m + RPNP Inscrito - RPNP Cancelado) / Rec. Corrente 12m

    A janela termina no ÚLTIMO MÊS COM DADO (D-1), fechado ou não — não depende
    mais da tabela mesfechado. A série exposta cobre os 12 últimos meses,
    identificados por (ano, mês), podendo cruzar o virador de exercício.
    """
    def _ano():
        return {m: 0.0 for m in range(1, 13)}

    rec      = {}   # rec[ano][mes]      receita corrente realizada
    desp_liq = {}   # desp_liq[ano][mes] despesas liquidadas correntes
    rpnp_can = {}   # rpnp_can[ano][mes] RPNP cancelado no exercício
    rpnp_ins = {}   # rpnp_ins[ano]      RPNP inscrito no encerramento (inmes=0)

    for r in rows:
        cc   = str(r.get("cocontacontabil") or "").strip()
        ccor = str(r.get("cocontacorrente") or "").strip()
        nat  = str(r.get("conatureza")      or "").strip()
        mes  = int(r.get("inmes") or 0)
        ano  = int(r.get("coexercicio") or 0)
        vacr = float(r.get("vacredito") or 0)
        vad  = float(r.get("vadebito")  or 0)
        try:
            cc_int = int(cc)
        except (ValueError, TypeError):
            continue
        nat2 = nat[1:2] if len(nat) >= 2 else ""
        val  = vacr - vad

        if (621200000 <= cc_int <= 621390199
                and ccor[:1] in ('1', '7')
                and 1 <= mes <= 12):
            rec.setdefault(ano, _ano())[mes] += val
        elif (cc_int in (622130300, 622130400)
                and nat2 in ('1', '2', '3')
                and 1 <= mes <= 12):
            desp_liq.setdefault(ano, _ano())[mes] += val
        elif (cc_int in (631100000, 631200000)
                and nat2 in ('1', '2', '3')
                and mes == 0):
            rpnp_ins[ano] = rpnp_ins.get(ano, 0.0) + val
        elif (cc_int == 631900000
                and nat2 in ('1', '2', '3', '7')
                and 1 <= mes <= 12):
            rpnp_can.setdefault(ano, _ano())[mes] += val

    def g(d, ano, mes):
        return d.get(ano, {}).get(mes, 0.0)

    # Mês de referência = último (ano, mês) com receita corrente movimentada.
    ref = None
    for ano, meses in rec.items():
        for mes, v in meses.items():
            if v != 0.0 and (ref is None or (ano, mes) > ref):
                ref = (ano, mes)
    if ref is None:
        agora = datetime.now()
        ref = (agora.year, agora.month)
    ref_ano, ref_mes = ref

    # Últimos 12 meses de referência (mais antigo -> mais recente).
    ref_meses = []
    yy, mm = ref_ano, ref_mes
    for _ in range(12):
        ref_meses.append((yy, mm))
        mm -= 1
        if mm == 0:
            mm, yy = 12, yy - 1
    ref_meses.reverse()

    serie = []
    for (y, m) in ref_meses:
        # Janela móvel de 12 meses terminando em (y, m): (y-1, m+1) .. (y, m).
        rec_12m  = (sum(g(rec, y - 1, k)      for k in range(m + 1, 13))
                  + sum(g(rec, y, k)          for k in range(1, m + 1)))
        desp_12m = (sum(g(desp_liq, y - 1, k) for k in range(m + 1, 13))
                  + sum(g(desp_liq, y, k)     for k in range(1, m + 1)))
        rpnp_i   = rpnp_ins.get(y, 0.0)                          # inscrito no encerramento do exerc. corrente
        rpnp_c   = sum(g(rpnp_can, y, k) for k in range(1, m + 1))  # cancelado no exerc. corrente
        desp_cor = desp_12m + rpnp_i - rpnp_c
        pct      = round(desp_cor / rec_12m * 100, 2) if rec_12m else None
        serie.append({
            "ano":                 y,
            "mes":                 m,
            "rec_corrente_12m":    round(rec_12m,  2),
            "desp_liquidadas_12m": round(desp_12m, 2),
            "rpnp_inscrito":       round(rpnp_i,   2),
            "rpnp_cancelado":      round(rpnp_c,   2),
            "desp_corrente_12m":   round(desp_cor, 2),
            "poupanca_pct":        pct,
        })

    rotulos = ["{}/{:02d}".format(s["ano"], s["mes"]) for s in serie]
    log.info("  Poupança Corrente: referência={}/{:02d}, série={}".format(
        ref_ano, ref_mes, rotulos))
    return {
        "ano_atual":  ref_ano,
        "ref_ano":    ref_ano,
        "ref_mes":    ref_mes,
        "limite_pct": 95.0,
        "serie":      serie,
    }


def save_poupanca_corrente_gz(D_obj):
    payload = {"atualizado_em": datetime.now(timezone.utc).isoformat()}
    payload.update(D_obj)
    content = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    gz_path = GZ_DIR / "poupanca_corrente.json.gz"
    with gzip.open(gz_path, "wb", compresslevel=9) as f:
        f.write(content)
    size_kb = gz_path.stat().st_size / 1024
    log.info(f"  poupanca_corrente.json.gz -- {size_kb:.1f} KB")


def upsert_poupanca_corrente_supabase(D_obj):
    """Envia poupanca_corrente para o Supabase como JSONB (1 linha por ano)."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        log.warning("  Supabase: nao configurado. Pulando upsert poupanca_corrente.")
        return
    try:
        ano = D_obj.get("ano_atual") or datetime.now().year
        payload = [{
            "ano":           ano,
            "dados":         D_obj,
            "atualizado_em": datetime.now(timezone.utc).isoformat(),
        }]
        total = _supabase_upsert("poupanca_corrente", payload, "ano")
        log.info(f"  Supabase: poupanca_corrente {ano} enviada ({total} linha).")
    except Exception as e:
        log.error(f"  Supabase poupanca_corrente falhou: {type(e).__name__}: {e}")


def build_resultado_primario_nominal_data(rows):
    """
    Computa o RREO Anexo 06 (Acima da Linha) com todos os sub-níveis,
    por mês individual. O browser soma os meses selecionados.
    Regras: tools/regra_ressultado_primario_nominal.txt
    """
    def _fmt3(arr):
        return {"a": round(arr[0], 2), "b": round(arr[1], 2), "c": round(arr[2], 2),
                "total": round(arr[0] + arr[1] + arr[2], 2)}

    REC_KEYS = [
        # Principais (já existiam)
        "I","II","III","V","VII","VIII","IX","X","XI","XII","XIV",
        # Sub-níveis correntes EXCETO RPPS
        "icms","ipva","itcd","iptu","itbi","iss","ir","outros_impostos","taxas",
        "contribuicoes",
        "outras_patrimoniais",
        "fpe","fpm","itr_trans","lc87","lc61","fundeb","outras_transf_corr",
        "correntes_restantes",
        # Sub-níveis capital EXCETO RPPS
        "outras_alienacoes","convenios","outras_transf_cap","outras_cap_prim",
    ]
    DEP_KEYS = [
        # Principais (já existiam)
        "XVIII","XIX","XXI","XXIII","XXIV","XXV","XXVI","XXVII","XXIX","XXX",
        # Sub-níveis
        "pessoal","outras_correntes_dep","investimentos","demais_inv",
    ]

    M_rec = {m: {k: 0.0 for k in REC_KEYS} for m in range(1, 13)}
    M_dep = {m: {k: [0.0, 0.0, 0.0] for k in DEP_KEYS} for m in range(1, 13)}
    M_jur = {m: {"XXXVI": 0.0, "XXXVII": 0.0} for m in range(1, 13)}
    P_rec = {k: 0.0 for k in REC_KEYS}

    _APLIC_FIN = {"132101","132102","132103","132104","132105","132999",
                  "732101","732102","732103","732104","732105","732999"}
    _FUNDEB_4  = {"1715","1751","7715","7751"}

    def _classifica_rec(val, co, cc_corr, is_exceto_rpps, bucket):
        co2, co3, co4, co6, co7 = co[:2], co[:3], co[:4], co[:6], co[:7]
        try:
            co_int = int(co)
        except (ValueError, TypeError):
            co_int = 0
        is_corrente = co2 in {"11","12","13","14","15","16","17","19",
                               "71","72","73","74","75","76","77","79"}
        is_capital  = co2 in {"21","22","23","24","29","81","82","83","84","89"}
        is_aplic    = cc_corr[:6] in _APLIC_FIN
        is_out_fin  = (co4 in {"1944","7944"} or
                       co6 in {"164101","164103","199911","764101","764103","799911"} or
                       co7 in {"1922012","1922064","1922142","1999993",
                                "7922012","7922064","7922142","7999993"})

        if is_exceto_rpps:
            if is_corrente:
                bucket["I"] += val
                if is_aplic:
                    bucket["II"] += val
                elif is_out_fin:
                    bucket["III"] += val
                # Sub-categorias
                if co2 in {"11","71"}:
                    if   (11145010 <= co_int <= 11145099) or (11145200 <= co_int <= 11145299):
                        bucket["icms"]          += val
                    elif co6 in {"111251","711251"}:
                        bucket["ipva"]          += val
                    elif 11125200 <= co_int <= 11125299:
                        bucket["itcd"]          += val
                    elif 11125000 <= co_int <= 11125099:
                        bucket["iptu"]          += val
                    elif 11125300 <= co_int <= 11125399:
                        bucket["itbi"]          += val
                    elif co7 in {"1114511","1114512","7114511","7114512"}:
                        bucket["iss"]           += val
                    elif co6 in {"111303","711303"}:
                        bucket["ir"]            += val
                    elif co4 in {"1119"}:
                        bucket["outros_impostos"] += val
                    elif co3 in {"112","712"}:
                        bucket["taxas"]         += val
                elif co2 in {"12","72"}:
                    bucket["contribuicoes"]     += val
                elif co2 in {"13","73"}:
                    if not is_aplic:
                        bucket["outras_patrimoniais"] += val
                elif co2 in {"17","77"}:
                    if   17115000 <= co_int <= 17115099:
                        bucket["fpe"]            += val
                    elif 17115100 <= co_int <= 17115199:
                        bucket["fpm"]            += val
                    elif 17115200 <= co_int <= 17115299:
                        bucket["itr_trans"]      += val
                    elif 17115300 <= co_int <= 17115399:
                        bucket["lc61"]           += val
                    elif co4 in _FUNDEB_4:
                        bucket["fundeb"]         += val
                    else:
                        bucket["outras_transf_corr"] += val
                elif co2 in {"14","15","16","19","74","75","76","79"}:
                    if not is_out_fin:
                        bucket["correntes_restantes"] += val
            elif is_capital:
                bucket["VII"] += val
                if   co2 in {"21","81"}:
                    bucket["VIII"] += val
                elif co2 in {"23","83"}:
                    bucket["IX"]   += val
                elif co6 in {"221101","821101"}:
                    bucket["X"]    += val
                elif co6 in {"221102","821102"}:
                    bucket["XI"]   += val
                elif co2 in {"22","82"}:
                    bucket["outras_alienacoes"] += val
                elif co2 in {"24","84"}:
                    if (co4 in {"2414","2422","2432","8414","8422","8432"} or
                            co6 in {"244150","244151","844150","844151"}):
                        bucket["convenios"]          += val
                    else:
                        bucket["outras_transf_cap"]  += val
                elif co3 in {"292","293","294","892","893","894"}:
                    bucket["XII"] += val
                elif co3 in {"291","299","891","899"}:
                    bucket["outras_cap_prim"] += val
        else:
            if is_corrente and not (is_aplic or is_out_fin):
                bucket["V"] += val
            elif ((co2 in {"22","24","82","84"} or co3 in {"291","299","891","899"}) and
                  co6 not in {"221101","221102","821101","821102"}):
                bucket["XIV"] += val

    for r in rows:
        cc = str(r.get("cocontacontabil") or "").strip()
        if not cc:
            continue
        mes     = int(r.get("inmes") or 0)
        vacred  = float(r.get("vacredito") or 0)
        vadeb   = float(r.get("vadebito")  or 0)
        cf      = str(r.get("cofontefederal") or "").strip()
        co      = str(r.get("coclasseorc")    or "").strip()
        cc_corr = str(r.get("cocontacorrente") or "").strip()
        na      = str(r.get("conatureza")     or "").strip()
        func    = str(r.get("cofuncao")       or "").strip()
        is_rpps        = cf[1:4] in {"800","801","802"} if len(cf) >= 4 else False
        is_exceto_rpps = not is_rpps
        cc4 = cc[:4]
        cc7 = cc[:7]

        if cc4 in {"6212","6213"}:
            if not mes or mes > 12:
                continue
            _classifica_rec(vacred - vadeb, co, cc_corr, is_exceto_rpps, M_rec[mes])

        elif cc4 in {"5211","5212"}:
            _classifica_rec(vadeb - vacred, co, cc_corr, is_exceto_rpps, P_rec)

        else:
            if not mes or mes > 12:
                continue
            if   cc7 == "6221304":                col, val = 0, vacred - vadeb
            elif cc4 == "6322":                   col, val = 1, vacred - vadeb
            elif cc in {"631400000","631820000"}:  col, val = 2, vacred - vadeb
            elif _is_xxxvi(cc)  and is_exceto_rpps:
                M_jur[mes]["XXXVI"]  += vacred - vadeb; continue
            elif _is_xxxvii(cc) and is_exceto_rpps:
                M_jur[mes]["XXXVII"] += vadeb - vacred; continue
            else:
                continue

            na2    = na[:2]
            na_mod = na[4:6]

            if func == "99":
                M_dep[mes]["XXIX"][col] += val
            elif is_exceto_rpps:
                if na2 in {"31","32","33"}:
                    M_dep[mes]["XVIII"][col] += val
                    if   na2 == "31": M_dep[mes]["pessoal"][col]              += val
                    elif na2 == "32": M_dep[mes]["XIX"][col]                  += val
                    elif na2 == "33": M_dep[mes]["outras_correntes_dep"][col] += val
                elif na2 in {"44","45","46"}:
                    M_dep[mes]["XXIII"][col] += val
                    if na2 == "44":
                        M_dep[mes]["investimentos"][col] += val
                    elif na2 == "45":
                        if   na_mod == "66": M_dep[mes]["XXIV"][col]     += val
                        elif na_mod == "64": M_dep[mes]["XXV"][col]      += val
                        elif na_mod == "63": M_dep[mes]["XXVI"][col]     += val
                        else:                M_dep[mes]["demais_inv"][col] += val
                    elif na2 == "46":
                        M_dep[mes]["XXVII"][col] += val
            else:
                if na2 in {"31","33"}:
                    M_dep[mes]["XXI"][col] += val
                elif na2 in {"44","45"} and not (na2 == "45" and na_mod in {"63","64","66"}):
                    M_dep[mes]["XXX"][col] += val

    max_mes = next((m for m in range(12, 0, -1)
                    if any(M_rec[m][k] != 0.0 for k in REC_KEYS)), 0)
    if max_mes == 0:
        max_mes = next((m for m in range(12, 0, -1)
                        if any(any(v != 0.0 for v in M_dep[m][k]) for k in DEP_KEYS)), 1)

    P_IV   = P_rec["I"]   - P_rec["II"]  - P_rec["III"]
    P_XIII = P_rec["VII"] - (P_rec["VIII"] + P_rec["IX"] + P_rec["X"] + P_rec["XI"] + P_rec["XII"])
    previsao = {k: round(P_rec[k], 2) for k in REC_KEYS}
    previsao.update({
        "IV":   round(P_IV, 2),
        "XIII": round(P_XIII, 2),
        "XVI":  round(P_IV + P_rec["V"] + P_XIII + P_rec["XIV"], 2),
        "XVII": round(P_IV + P_XIII, 2),
    })

    por_mes = {}
    for mes in range(1, max_mes + 1):
        por_mes[str(mes)] = {
            "rec": {k: round(M_rec[mes][k], 2) for k in REC_KEYS},
            "dep": {k: [round(M_dep[mes][k][i], 2) for i in range(3)] for k in DEP_KEYS},
            "jur": {"XXXVI": round(M_jur[mes]["XXXVI"], 2),
                    "XXXVII": round(M_jur[mes]["XXXVII"], 2)},
        }

    log.info(f"  Resultado Primário/Nominal: max_mes={max_mes}, meses={list(por_mes.keys())}")
    return {"max_mes": max_mes, "previsao": previsao, "por_mes": por_mes}


def build_minimo_saude_data(rows):
    """
    Computa o RREO Anexo 08 (Mínimo de Saúde/ASPS) organizado por bimestre acumulado.
    Regras: tools/regra_minimo_saude.txt (LC 141/2012)
    """
    ano_atual = datetime.now().year

    SUBFUNCOES_ASPS = {301, 302, 303, 304, 305, 306}
    OUTRAS_SUB      = {122, 126, 128, 364}
    SUBF_MAP        = {301: "IV", 302: "V", 303: "VI", 304: "VII", 305: "VIII", 306: "IX"}
    DEP_KEYS        = ["IV", "V", "VI", "VII", "VIII", "IX", "X"]
    REC_KEYS        = [
        "icms_75", "multas_icms_75", "fcp_75", "itcd", "ipva_est", "irrf",
        "fpe", "ipi_75", "outras_transf_75",
        "icms_25", "multas_icms_25", "fcp_25", "ipva_mun", "iptu", "itbi", "iss",
        "fpm", "itr", "ipi_25", "outras_transf_25",
    ]
    I_KEYS  = REC_KEYS[:9]
    II_KEYS = REC_KEYS[9:]

    def _acum_rec(dest, c8, saldo):
        c = c8.ljust(8, "0")
        if   c == "11145011":            dest["icms_75"] += saldo * 0.75;  dest["icms_25"] += saldo * 0.25
        elif "11145013" <= c <= "11145018":dest["multas_icms_75"] += saldo * 0.75; dest["multas_icms_25"] += saldo * 0.25
        elif "11145021" <= c <= "11145026":dest["fcp_75"] += saldo * 0.75;   dest["fcp_25"] += saldo * 0.25
        elif c == "11125201":            dest["itcd"] += saldo
        elif "11125203" <= c <= "11125208": dest["itcd"] += saldo
        elif c == "11125101":            dest["ipva_est"] += saldo * 0.50; dest["ipva_mun"] += saldo * 0.50
        elif "11125103" <= c <= "11125108": dest["ipva_est"] += saldo * 0.50; dest["ipva_mun"] += saldo * 0.50
        elif "11130201" <= c <= "11130341": dest["irrf"] += saldo
        elif c == "17115001":            dest["fpe"] += saldo
        elif c == "17115301":            dest["ipi_75"] += saldo * 0.75;   dest["ipi_25"] += saldo * 0.25
        elif c == "17196201":            dest["outras_transf_75"] += saldo * 0.75; dest["outras_transf_25"] += saldo * 0.25
        elif c == "11125001":            dest["iptu"] += saldo
        elif "11125003" <= c <= "11125008": dest["iptu"] += saldo
        elif c == "11125301":            dest["itbi"] += saldo
        elif "11125303" <= c <= "11125308": dest["itbi"] += saldo
        elif c == "11145111":            dest["iss"] += saldo
        elif "11145113" <= c <= "11145118": dest["iss"] += saldo
        elif c == "17115111":            dest["fpm"] += saldo
        elif c == "17115201":            dest["itr"] += saldo

    def _dep_key(func_r, subf_r):
        try:
            func_i = int(func_r); subf_i = int(subf_r)
        except (ValueError, TypeError):
            return None
        if func_i != 10:
            return None
        if subf_i in SUBFUNCOES_ASPS:
            return SUBF_MAP[subf_i]
        if subf_i in OUTRAS_SUB:
            return "X"
        return None

    # Receitas: previsão inicial e atualizada (atemporais)
    rec_prev_ini = {k: 0.0 for k in REC_KEYS}
    rec_prev_atu = {k: 0.0 for k in REC_KEYS}
    # Receitas realizadas por mês
    rec_real = {m: {k: 0.0 for k in REC_KEYS} for m in range(1, 13)}
    # Dotações (anuais, não mensais) por key de despesa
    dot_ini = {k: {"cor": 0.0, "cap": 0.0} for k in DEP_KEYS}
    dot_atu = {k: {"cor": 0.0, "cap": 0.0} for k in DEP_KEYS}
    # Execução por mês e key de despesa
    exec_dep = {m: {k: {"cor": {"emp": 0.0, "liq": 0.0, "pago": 0.0},
                         "cap": {"emp": 0.0, "liq": 0.0, "pago": 0.0}}
                    for k in DEP_KEYS}
                for m in range(1, 13)}

    max_mes_real = 0

    for r in rows:
        cc_raw = str(r.get("cocontacontabil") or "").strip()
        ccor   = str(r.get("cocontacorrente") or "").strip()
        nat    = str(r.get("conatureza")      or "").strip()
        func_r = str(r.get("cofuncao")        or "").strip()
        subf_r = str(r.get("cosubfuncao")     or "").strip()
        mes_r  = int(r.get("inmes")           or 0)
        vad    = float(r.get("vadebito")      or 0)
        vacr   = float(r.get("vacredito")     or 0)

        try:
            cc_int = int(cc_raw)
        except (ValueError, TypeError):
            continue

        cc4 = cc_raw[:4]
        cc5 = cc_raw[:5]
        cc7 = cc_raw[:7]
        c8  = ccor[:8]

        # Receita: previsão inicial (521110000 conta para ini E atu)
        if cc_int == 521110000:
            _acum_rec(rec_prev_ini, c8, vad - vacr)
        # Receita: previsão atualizada
        if cc_int in (521110000, 521210100, 5212102002):
            _acum_rec(rec_prev_atu, c8, vad - vacr)
            continue

        # Receita: realizado
        if cc_int in (621200000, 621300000):
            if 1 <= mes_r <= 12:
                _acum_rec(rec_real[mes_r], c8, vacr - vad)
                if mes_r > max_mes_real:
                    max_mes_real = mes_r
            continue

        # Classificar despesa (COFUNCAO=10, COSUBFUNCAO adequado)
        key = _dep_key(func_r, subf_r)
        if key is None:
            continue

        nat2 = nat[1:2]  # SUBSTR(CONATUREZA,2,1)
        is_cor = nat2 in ("1", "3")
        is_cap = nat2 in ("4", "6")

        # Dotação: cc5=52211 → ini+atu; cc5 in 52212/52215/52219 → só atu
        if cc5 in ("52211", "52212", "52215", "52219"):
            saldo = vad - vacr
            if cc5 == "52211":
                if   is_cor: dot_ini[key]["cor"] += saldo
                elif is_cap: dot_ini[key]["cap"] += saldo
            if   is_cor: dot_atu[key]["cor"] += saldo
            elif is_cap: dot_atu[key]["cap"] += saldo
            continue

        # Execução: cc5=62213 → empenhada; cc7 sub-códigos → liquidada e paga
        if cc5 == "62213" and 1 <= mes_r <= 12:
            saldo = vacr - vad
            if   is_cor: exec_dep[mes_r][key]["cor"]["emp"] += saldo
            elif is_cap: exec_dep[mes_r][key]["cap"]["emp"] += saldo
            if cc7 in ("6221303", "6221304", "6221307"):
                if   is_cor: exec_dep[mes_r][key]["cor"]["liq"] += saldo
                elif is_cap: exec_dep[mes_r][key]["cap"]["liq"] += saldo
            if cc7 == "6221304":
                if   is_cor: exec_dep[mes_r][key]["cor"]["pago"] += saldo
                elif is_cap: exec_dep[mes_r][key]["cap"]["pago"] += saldo

    # Determinar max_bimestre
    max_bimestre = (max_mes_real + 1) // 2 if max_mes_real else 0
    if max_bimestre == 0:
        log.warning("  Mínimo Saúde: nenhum dado de receita realizada encontrado.")
        return {"ano": ano_atual, "max_bimestre": 0, "por_bimestre": {}}

    BIM_FIM = {1: 2, 2: 4, 3: 6, 4: 8, 5: 10, 6: 12}

    por_bimestre = {}
    for bim in range(1, max_bimestre + 1):
        mes_fim   = min(BIM_FIM[bim], max_mes_real)
        is_ultimo = (bim == 6)

        # Receitas acumuladas até mes_fim
        rec_acum = {k: sum(rec_real[m][k] for m in range(1, mes_fim + 1)) for k in REC_KEYS}
        I_pi  = sum(rec_prev_ini[k] for k in I_KEYS)
        I_pa  = sum(rec_prev_atu[k] for k in I_KEYS)
        I_r   = sum(rec_acum[k]     for k in I_KEYS)
        II_pi = sum(rec_prev_ini[k] for k in II_KEYS)
        II_pa = sum(rec_prev_atu[k] for k in II_KEYS)
        II_r  = sum(rec_acum[k]     for k in II_KEYS)
        III_r = I_r + II_r

        # Execução acumulada por key
        exec_acum = {k: {"cor": {"emp": 0.0, "liq": 0.0, "pago": 0.0},
                         "cap": {"emp": 0.0, "liq": 0.0, "pago": 0.0}}
                     for k in DEP_KEYS}
        for m in range(1, mes_fim + 1):
            for k in DEP_KEYS:
                for cat in ("cor", "cap"):
                    for col in ("emp", "liq", "pago"):
                        exec_acum[k][cat][col] += exec_dep[m][k][cat][col]

        def _dep_obj(k):
            c = exec_acum[k]["cor"]; p = exec_acum[k]["cap"]
            di     = dot_ini[k]["cor"] + dot_ini[k]["cap"]
            da     = dot_atu[k]["cor"] + dot_atu[k]["cap"]
            emp_k  = c["emp"]  + p["emp"]
            liq_k  = c["liq"]  + p["liq"]
            pago_k = c["pago"] + p["pago"]
            rpnp_k = max(0.0, emp_k - liq_k) if is_ultimo else 0.0
            return {
                "dot_ini": round(di,     2),
                "dot_atu": round(da,     2),
                "emp":     round(emp_k,  2),
                "liq":     round(liq_k,  2),
                "pago":    round(pago_k, 2),
                "rpnp":    round(rpnp_k, 2),
                "cor": {"dot_ini": round(dot_ini[k]["cor"], 2),
                        "dot_atu": round(dot_atu[k]["cor"], 2),
                        "emp":  round(c["emp"],  2),
                        "liq":  round(c["liq"],  2),
                        "pago": round(c["pago"], 2)},
                "cap": {"dot_ini": round(dot_ini[k]["cap"], 2),
                        "dot_atu": round(dot_atu[k]["cap"], 2),
                        "emp":  round(p["emp"],  2),
                        "liq":  round(p["liq"],  2),
                        "pago": round(p["pago"], 2)},
            }

        XI_di   = sum(dot_ini[k]["cor"] + dot_ini[k]["cap"] for k in DEP_KEYS)
        XI_da   = sum(dot_atu[k]["cor"] + dot_atu[k]["cap"] for k in DEP_KEYS)
        XI_emp  = sum(exec_acum[k]["cor"]["emp"]  + exec_acum[k]["cap"]["emp"]  for k in DEP_KEYS)
        XI_liq  = sum(exec_acum[k]["cor"]["liq"]  + exec_acum[k]["cap"]["liq"]  for k in DEP_KEYS)
        XI_pago = sum(exec_acum[k]["cor"]["pago"] + exec_acum[k]["cap"]["pago"] for k in DEP_KEYS)
        XI_rpnp = max(0.0, XI_emp - XI_liq) if is_ultimo else 0.0

        XVI_emp   = XI_emp
        XVI_liq   = XI_liq
        XVII      = round(I_r * 0.12 + II_r * 0.15, 2)
        XVIII_emp = round(XVI_emp - XVII, 2)
        XVIII_liq = round(XVI_liq - XVII, 2)
        pct_liq   = round(XVI_liq / III_r * 100, 2) if III_r else None
        pct_emp   = round(XVI_emp / III_r * 100, 2) if III_r else None
        pct_min   = round(XVII    / III_r * 100, 2) if III_r else None

        def _det(k):
            return {"prev_ini": round(rec_prev_ini[k], 2),
                    "prev_atu": round(rec_prev_atu[k], 2),
                    "realizado": round(rec_acum[k], 2)}

        por_bimestre[str(bim)] = {
            "is_ultimo": is_ultimo,
            "rec": {
                "I":   {"prev_ini": round(I_pi, 2),  "prev_atu": round(I_pa, 2),  "realizado": round(I_r, 2)},
                "II":  {"prev_ini": round(II_pi, 2), "prev_atu": round(II_pa, 2), "realizado": round(II_r, 2)},
                "III": {"prev_ini": round(I_pi + II_pi, 2),
                        "prev_atu": round(I_pa + II_pa, 2),
                        "realizado": round(III_r, 2)},
                "detalhe": {k: _det(k) for k in REC_KEYS},
            },
            "dep": {
                "IV":  _dep_obj("IV"),  "V":   _dep_obj("V"),
                "VI":  _dep_obj("VI"),  "VII": _dep_obj("VII"),
                "VIII":_dep_obj("VIII"),"IX":  _dep_obj("IX"),
                "X":   _dep_obj("X"),
                "XI": {"dot_ini": round(XI_di,   2), "dot_atu": round(XI_da,   2),
                       "emp":     round(XI_emp,  2), "liq":     round(XI_liq,  2),
                       "pago":    round(XI_pago, 2), "rpnp":    round(XI_rpnp, 2)},
            },
            "apuracao": {
                "XII":   {"emp": round(XI_emp,  2), "liq": round(XI_liq,  2), "pago": round(XI_pago, 2)},
                "XIII":  0.0, "XIV": 0.0, "XV": 0.0,
                "XVI":   {"emp": round(XVI_emp, 2), "liq": round(XVI_liq, 2), "pago": round(XI_pago, 2)},
                "XVII":  XVII,
                "XVIII": {"emp": round(XVIII_emp, 2), "liq": round(XVIII_liq, 2)},
                "pct_liq": pct_liq, "pct_emp": pct_emp, "pct_minimo": pct_min,
            },
        }

    log.info(f"  Mínimo Saúde: max_bimestre={max_bimestre}, max_mes_real={max_mes_real}")
    return {"ano": ano_atual, "max_bimestre": max_bimestre, "por_bimestre": por_bimestre}


def save_minimo_saude_gz(D_obj):
    payload = {"atualizado_em": datetime.now(timezone.utc).isoformat()}
    payload.update(D_obj)
    content = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    gz_path = GZ_DIR / "minimo_saude.json.gz"
    with gzip.open(gz_path, "wb", compresslevel=9) as f:
        f.write(content)
    size_kb = gz_path.stat().st_size / 1024
    log.info(f"  minimo_saude.json.gz -- {size_kb:.1f} KB")


def upsert_minimo_saude_supabase(D_obj):
    if not SUPABASE_URL or not SUPABASE_KEY:
        log.warning("  Supabase: nao configurado. Pulando upsert minimo_saude.")
        return
    try:
        ano = D_obj.get("ano") or datetime.now().year
        payload = [{
            "ano":           ano,
            "dados":         D_obj,
            "atualizado_em": datetime.now(timezone.utc).isoformat(),
        }]
        total = _supabase_upsert("minimo_saude", payload, "ano")
        log.info(f"  Supabase: minimo_saude {ano} enviada ({total} linha).")
    except Exception as e:
        log.error(f"  Supabase minimo_saude falhou: {type(e).__name__}: {e}")


def build_minimo_educacao_data(rows):
    """
    Computa o RREO MDE (Mínimo da Educação) organizado por mês acumulado.
    Quadros: Receita Resultante de Impostos (Arts. 212 e 212-A CF) + FUNDEB.
    Regras: tools/regra_minimo_educacao.txt
    """
    ano_atual = datetime.now().year
    ano_ant   = ano_atual - 1

    # L8.1 / L8.2 = SUPERÁVIT DO EXERCÍCIO IMEDIATAMENTE ANTERIOR / RESIDUAL
    #   mil2001.saldocontabil_ex não retém dados de ano_atual-2, então L8 não é
    #   calculável automaticamente — informado manualmente, a partir do L8 (8.1+8.2)
    #   publicado no RREO. É re-apurado ao longo do ano conforme o fechamento do
    #   balanço do exercício anterior, então cada BIMESTRE usa o valor vigente no
    #   seu próprio RREO. MDE_SUPERAVIT_8_1_POR_BIM mapeia bim→valor; bimestres
    #   sem entrada herdam o último valor informado (fill-forward).
    #   2026: 31.361.102,40 (1º bim) → 32.347.113,94 (a partir do 2º bim).
    MDE_SUPERAVIT_8_1_POR_BIM = {1: 31361102.40, 2: 32347113.94}
    MDE_SUPERAVIT_8_2 = 0.00          # 8.2 — superávit residual de outros exercícios

    def _l81_por_bim():
        out, last = {}, 0.0
        for bim in range(1, 7):
            if bim in MDE_SUPERAVIT_8_1_POR_BIM:
                last = MDE_SUPERAVIT_8_1_POR_BIM[bim]
            out[str(bim)] = round(last, 2)
        return out

    # Acumulador de receita FUNDEB do ano anterior para L19(s)
    # L19(s) = 10% × receita FUNDEB do ano anterior
    #   f61    : NR=17515001 + fonte_mae=122  (19.1)
    #   f_1715 : SUBSTR(ccor,1,4)='1715' (COED=1715)  (19.2)
    fundeb_ant = {"f61": 0.0, "f_1715": 0.0}

    REC_KEYS = [
        "icms_principal", "icms_adicional",
        "itcd", "ipva", "iptu", "itbi", "iss", "irrf", "outros_impostos",
        "fpe", "fpm", "ipi_exp", "itr", "iof_ouro", "outras_transf",
    ]
    FUNDEB_KEYS = [
        "f61_principal", "f61_rend_aplic", "f61_ressarc",
        "f62_principal", "f62_rend_aplic", "f62_ressarc",
        "f63_principal", "f63_rend_aplic", "f63_ressarc",
        "f64_principal", "f64_rend_aplic", "f64_ressarc",
    ]
    REND_NR6    = {"132101", "132102", "132103", "132105", "132999"}
    RESSARC_NR6 = {"192251"}

    rec_prev   = {k: 0.0 for k in REC_KEYS}
    fund_prev  = {k: 0.0 for k in FUNDEB_KEYS}
    rec_real   = {m: {k: 0.0 for k in REC_KEYS}    for m in range(1, 13)}
    fund_real  = {m: {k: 0.0 for k in FUNDEB_KEYS} for m in range(1, 13)}
    max_mes    = 0

    # ── Despesa MDE ──────────────────────────────────────────────────────────
    # Subfunções do Quadro 10 (FUNDEB): inclui 122=Adm.Geral e transp=781-785
    SUBFUNC_Q10    = {"361","362","363","365","366","367","122"}
    SUBFUNC_TRANSP = {"781","782","783","784","785"}          # Transporte Escolar (10.2.7)
    SUBFUNC_KEYS   = sorted(SUBFUNC_Q10) + ["transp", "outras"]
    # FR suffix = últimos 3 dígitos do COFONTEFEDERAL; X540-X543 + X546 = FUNDEB
    FR_FUNDEB      = {"540","541","542","543","546"}
    # NDs excluídas do grupo Profissionais mesmo sendo 31xxx (encargos patronais)
    ND_EXCL_PROF   = {"31900100","31900300"}

    def _sub_bucket():
        """Acumuladores por subfunção: dot_atu, emp, liq, pago."""
        return {s: {"dot_atu": 0.0, "emp": 0.0, "liq": 0.0, "pago": 0.0}
                for s in SUBFUNC_KEYS}

    # CO=1001: plano (sem split prof/outras — Quadro 20 vem depois)
    desp_dot_1001  = _sub_bucket()
    desp_real_1001 = {m: _sub_bucket() for m in range(1, 13)}

    # Quadro 10: split por grupo baseado em FR (fonte FUNDEB) + ND
    #   'prof'  = FR fundeb AND ND 31xxx exceto ND_EXCL_PROF
    #   'outras'= FR fundeb AND (ND não é 31xxx, OU ND em ND_EXCL_PROF)
    GRUPOS_1070 = ("prof", "outras")
    desp_dot_1070  = {g: _sub_bucket() for g in GRUPOS_1070}
    desp_real_1070 = {m: {g: _sub_bucket() for g in GRUPOS_1070} for m in range(1, 13)}

    # L19 (u)/(v): superávit do FUNDEB APLICADO no exercício, por mês.
    # = despesa FUNDEB (Quadro 10) custeada com recursos de exercícios anteriores
    # (cofonte prefixo '3' = superávit financeiro), empenhada. (u) é a parcela
    # aplicada até o 1º quadrimestre (meses 1-4); (v) a parcela após. O dashboard
    # agrega por período e deriva (w)=max(0,t-u-v), (x)=max(0,t-u), L25=(x).
    superavit_aplicado = {m: 0.0 for m in range(1, 13)}

    # Quadro 20: Despesas MDE custeadas com Receitas de Impostos (exceto FUNDEB)
    # FR: 500 | 502 | 718  +  CO: 1001  +  COFUNCAO: 12
    # ND: grupo 3 (correntes) + grupo 4 (capital), exceto ND_EXCL_PROF (31900100, 31900300)
    FR_Q20       = {"500", "502", "718"}
    Q20_SUBS     = {"361","362","365","366","367","122"}  # linhas individuais
    Q20_TRANSP   = "785"                                  # 20.7 Transporte

    # Exclusões oficiais do Quadro 20 por DESDOBRAMENTO de ND (8 dígitos) —
    # aba "Exclusões" da planilha oficial (MDE20261B.xlsx), conferidas ao
    # centavo contra o RREO Anexo 8 do 1º bim/2026. Como a dotação no SIGGO
    # só existe em natureza de 6 dígitos, o RREO oficial ABATE da DOTAÇÃO
    # ATUALIZADA o EMPENHADO acumulado dessas fatias (e as exclui de
    # emp/liq/pago). O abate vale para todo o universo CO=1001 (fonte
    # resumida), SEM o filtro FR_Q20 — há fatias em fontes 1021/1002, de
    # COFONTEFEDERAL 1540 (FUNDEB), que o oficial abate mesmo assim. Não
    # restringir a lista de fontes: o critério oficial é a fonte resumida.
    ND8_EXCL_MDE = {"31901195", "31901131", "31901125", "33903023"}

    def _q20_bucket():
        keys = sorted(Q20_SUBS) + ["transp", "outras"]
        return {s: {"dot_atu":0.0, "emp":0.0, "liq":0.0, "pago":0.0} for s in keys}

    desp_real_q20 = {m: _q20_bucket() for m in range(1, 13)}

    # Quadro 21: Despesas MDE custeadas com Receitas de Impostos E FUNDEB
    # (FR 500/502/718; CO 1001) + (FR 540/541/542/543/546), cofuncao=12
    # 21.1 (subfunção 365) é dividido em creche (21.1.1) e pré-escola (21.1.2)
    # via COPROJETO=2388 + COSUBTITULO (1=creche, 2=pré-escola); demais
    # combinações com subfunção 365 caem no resíduo "365" (educação infantil geral)
    Q21_KEYS = ["365", "361", "362", "364", "demais"]

    def _q21_bucket():
        return {s: {"dot_atu": 0.0, "emp": 0.0, "liq": 0.0, "pago": 0.0} for s in Q21_KEYS}

    def _q21_key(subfunc, coprojeto=None, cosubtitulo=None):
        return subfunc if subfunc in ("365", "361", "362", "364") else "demais"

    desp_real_q21 = {m: _q21_bucket() for m in range(1, 13)}

    # Quadro 30: Restos a Pagar inscritos em exercícios anteriores — MDE
    #   30.1: FR 500/502/718, CO=1001 | 30.2: FR=540 | 30.3: FR 541/542/543
    #   colunas: ac=saldo inicial, ad=liquidados, ae=pagos, af=cancelados (ag = ac-ae-af)
    # Regra atualizada (2026-06-12): contas exatas (sem SUBSTR), soma direta
    # do saldo (convenção ímpar=deb-cred / par=cred-deb). Validado EXATAMENTE
    # (diff=0,00) contra RREO oficial 1º bim/2026 (30.1/30.2/30.3):
    #   ac = soma(CC9_AC),        inmes=0
    #   ae = soma(CC9_AE),        inmes 1-12
    #   ad = ae + soma(CC9_AD_EXTRA), inmes 1-12
    #   af = soma(CC9_AF),        inmes 1-12
    # Ver tools/regra_minimo_educacao.txt.
    Q30_KEYS   = ("30.1", "30.2", "30.3")
    FR_Q30_1   = {"500", "502", "718"}
    FR_Q30_2   = {"540"}
    FR_Q30_3   = {"541", "542", "543"}
    CC9_AC       = {"531100000", "531200000", "532100000", "532200000"}
    CC9_AE       = {"631400000", "631820000", "632210100", "632210200",
                     "632210300", "632210400"}
    CC9_AD_EXTRA = {"631300000", "631810000"}
    CC9_AF       = {"631900000", "632900000"}

    def _q30_bucket():
        return {k: {"ac": 0.0, "ae": 0.0, "ad_extra": 0.0, "af": 0.0} for k in Q30_KEYS}

    # Por mês, como os demais quadros — o seletor de período do dashboard
    # agrega os meses e calcula ag = ac - ae - af sobre o agregado.
    q30_acc = {m: _q30_bucket() for m in range(1, 13)}

    def _q30_grp(co_row, fr_suf):
        if co_row == "1001" and fr_suf in FR_Q30_1:
            return "30.1"
        if fr_suf in FR_Q30_2:
            return "30.2"
        if fr_suf in FR_Q30_3:
            return "30.3"
        return None

    # Indicadores FUNDEB (linhas 11-14) — sem dimensão de subfunção, só emp/liq/pago
    # l11_540=impostos, l11_541=VAAF, l11_542=VAAT, l11_543=VAAR
    # l12=profissionais(FR 540+541+542), l13=VAAT infantil, l14=VAAT capital
    IND_KEYS = ("l11_540","l11_541","l11_542","l11_543","l12","l13","l14")
    ind_real = {m: {k: {"emp":0.0,"liq":0.0,"pago":0.0} for k in IND_KEYS}
                for m in range(1, 13)}

    def classify_rec(nr):
        if nr[:7] == "1114501":                                            return "icms_principal"
        if nr[:7] == "1114502":                                            return "icms_adicional"
        if nr[:6] == "111252":                                             return "itcd"
        if nr[:6] == "111251":                                             return "ipva"
        if nr[:6] == "111250":                                             return "iptu"
        if nr[:6] == "111253":                                             return "itbi"
        if nr[:7] in ("1114511", "1114512"):                               return "iss"
        if nr[:6] == "111303":                                             return "irrf"
        if nr[:6] == "111999":                                             return "outros_impostos"
        if nr[:6] == "171150":                                             return "fpe"
        if nr[:7] in ("1711511", "1711512"):                               return "fpm"
        if nr[:6] in ("171153",  "172152"):                                return "ipi_exp"
        if nr[:6] == "171152":                                             return "itr"
        if nr[:6] == "171155":                                             return "iof_ouro"
        if nr[:6] in ("171961",  "171962", "171963", "172953"):            return "outras_transf"
        return None

    def classify_fundeb(nr, fr_suf):
        is_rend    = nr[:6] in REND_NR6
        is_ressarc = nr[:6] in RESSARC_NR6
        if fr_suf == "540":
            if nr[:6] == "175150":                 return "f61_principal"
            if is_rend:                             return "f61_rend_aplic"
            if is_ressarc:                          return "f61_ressarc"
        elif fr_suf in ("541", "546"):
            if nr[:6] in ("171551", "171553"):     return "f62_principal"
            if is_rend:                             return "f62_rend_aplic"
            if is_ressarc:                          return "f62_ressarc"
        elif fr_suf == "542":
            if nr[:6] == "171550":                 return "f63_principal"
            if is_rend:                             return "f63_rend_aplic"
            if is_ressarc:                          return "f63_ressarc"
        elif fr_suf == "543":
            if nr[:6] == "171552":                 return "f64_principal"
            if is_rend:                             return "f64_rend_aplic"
            if is_ressarc:                          return "f64_ressarc"
        return None

    def _is_superavit_fundeb(r, co_row, fr_suf):
        """Despesa da unidade do FUNDEB (couo 18903) custeada com superávit
        (cofonte de exercícios anteriores, prefixo '3') cujo COFONTEFEDERAL não
        é 54x. O CASE do SQL a rotula CO=1001 (casa antes do ramo 1070), mas o
        RREO oficial a classifica no FUNDEB (Quadro 10), não no Quadro 20. É o
        único caso em que couo 18903 + cofonte 3xx tem fr_suf fora de FR_FUNDEB
        (validado contra o Anexo 8 do 2º bim/2026 — linhas 10/20/21)."""
        return (co_row == "1001"
                and str(r.get("couo") or "").strip() == "18903"
                and str(r.get("cofonte") or "")[:1] == "3"
                and fr_suf not in FR_FUNDEB)

    for r in rows:
        # ── Separar ano corrente vs ano anterior ─────────────────────────────
        inano = int(r.get("inano") or ano_atual)

        cc_raw = str(r.get("cocontacontabil") or "").strip()
        ccor   = str(r.get("cocontacorrente") or "").strip()
        fr_raw = str(r.get("cofontefederal")  or "").strip()
        mes_v  = r.get("inmes")
        mes    = int(mes_v) if mes_v is not None else 0
        vad    = float(r.get("vadebito")  or 0)
        vacr   = float(r.get("vacredito") or 0)

        if not cc_raw:
            continue

        # Natureza: ímpar = devedora (vad-vacr); par = credora (vacr-vad)
        saldo = (vad - vacr) if cc_raw[0] in "13579" else (vacr - vad)
        if not saldo:
            continue

        # ── Ano anterior: acumula receita FUNDEB para L8.1/L19(s) ──────────────
        if inano == ano_ant:
            nr_ant      = ccor[:8] if len(ccor) >= 8 else ccor.ljust(8, "0")
            cofonte_ant = str(r.get("cofonte") or "").strip()

            # Contas EXATAS conforme a regra do L19(s) — não usar a família
            # 6212/6213: o SQL passou a trazer também contas de dedução
            # (ex.: 621390101) para a linha 1.8 do ano corrente, e elas NÃO
            # entram no cálculo do superávit (validado contra o Excel).
            if cc_raw in ("621200000", "621300000"):
                # L8.1/L19.1(s): NR=17515001 + fonte_mae=122
                if nr_ant == "17515001":
                    fundeb_ant["f61"] += saldo
                elif cofonte_ant[:3] == "122":
                    fundeb_ant["f61"] += saldo
                # L19.2(s): COED=1715 = SUBSTR(ccor,1,4)='1715'
                if nr_ant[:4] == "1715":
                    fundeb_ant["f_1715"] += saldo

            continue  # linhas do ano anterior não alimentam os quadros do ano corrente

        cc4    = cc_raw[:4]
        cc5    = cc_raw[:5]
        nr     = ccor[:8] if len(ccor) >= 8 else ccor.ljust(8, "0")
        fr_suf = fr_raw[-3:] if len(fr_raw) >= 3 else fr_raw

        if cc4 in ("5211", "5212"):
            # 521120101 = dedução FUNDEB — excluir da previsão
            if cc_raw == "521120101":
                continue
            # Previsão Atualizada (sem dimensão de mês)
            key = classify_rec(nr)
            if key:
                rec_prev[key] += saldo
            key_f = classify_fundeb(nr, fr_suf)
            if key_f:
                fund_prev[key_f] += saldo

        elif cc4 in ("6212", "6213"):
            if not (1 <= mes <= 12):
                continue
            # 621310100 = dedução de FUNDEB — excluir das realizadas
            if cc_raw == "621310100":
                continue
            # Receita realizada
            key = classify_rec(nr)
            if key:
                rec_real[mes][key] += saldo
                if mes > max_mes:
                    max_mes = mes
            # FUNDEB recebido (subconjunto filtrado por FR)
            key_f = classify_fundeb(nr, fr_suf)
            if key_f:
                fund_real[mes][key_f] += saldo

        elif cc5 in ("52211", "52212", "52215", "52219"):
            # Dotação MDE — acumulada por mês (igual à execução) para que o
            # seletor de período do dashboard filtre corretamente
            if not (1 <= mes <= 12):
                continue
            co_row  = str(r.get("co")          or "").strip()
            nd_row  = str(r.get("nd")          or "").strip()
            subfunc = str(r.get("cosubfuncao") or "").strip()
            cofunc  = str(r.get("cofuncao")    or "").strip()
            nd_ok   = nd_row[:8] not in ND_EXCL_PROF
            sup_fdb = _is_superavit_fundeb(r, co_row, fr_suf)

            if subfunc in SUBFUNC_TRANSP:
                sub_key = "transp"
            elif subfunc in SUBFUNC_Q10:
                sub_key = subfunc
            else:
                sub_key = "outras"

            # FR fundeb tem prioridade: CO=1001 com FR=540 ainda é Quadro 10.
            # sup_fdb: FUNDEB-superávit da UO 18903 que o CASE rotulou 1001 (ver
            # _is_superavit_fundeb) — o RREO oficial o conta no Quadro 10.
            if fr_suf in FR_FUNDEB or co_row == "1070" or sup_fdb:
                is_prof = nd_row[:2] == "31" and nd_row[:8] not in ND_EXCL_PROF
                grupo   = "prof" if is_prof else "outras"
                desp_real_1070[mes][grupo][sub_key]["dot_atu"] += saldo
            elif co_row == "1001":
                desp_real_1001[mes][sub_key]["dot_atu"] += saldo

            # Quadro 20: CO=1001, FR 500/502/718, nd_ok (exceto FUNDEB-superávit)
            #   cofuncao=12 → subfunções individuais + catchall "outras"
            #   cofuncao=28 + cosubfuncao 843/844 → "outras" (mesmo que 10.2.8)
            if co_row == "1001" and fr_suf in FR_Q20 and nd_ok and not sup_fdb:
                if cofunc == "12":
                    q20k = "transp" if subfunc == Q20_TRANSP \
                           else (subfunc if subfunc in Q20_SUBS else "outras")
                elif cofunc == "28" and subfunc in ("843","844"):
                    q20k = "outras"
                else:
                    q20k = None
                if q20k:
                    desp_real_q20[mes][q20k]["dot_atu"] += saldo

            # Quadro 21: (CO=1001 e FR 500/502/718) OU FR FUNDEB OU FUNDEB-superávit
            if cofunc == "12" and nd_ok and \
               ((co_row == "1001" and fr_suf in FR_Q20) or fr_suf in FR_FUNDEB
                or sup_fdb):
                coproj_q21 = str(r.get("coprojeto")   or "").strip()
                cosubt_q21 = str(r.get("cosubtitulo") or "").strip()
                q21k = _q21_key(subfunc, coproj_q21, cosubt_q21)
                if q21k:
                    desp_real_q21[mes][q21k]["dot_atu"] += saldo

        elif cc5 == "62213":
            # Despesas executadas MDE (empenhadas / liquidadas / pagas)
            if not (1 <= mes <= 12):
                continue
            co_row  = str(r.get("co") or "").strip()
            nd_row  = str(r.get("nd") or "").strip()
            subfunc = str(r.get("cosubfuncao") or "").strip()
            if subfunc in SUBFUNC_TRANSP:
                sub_key = "transp"
            elif subfunc in SUBFUNC_Q10:
                sub_key = subfunc
            else:
                sub_key = "outras"
            cc7 = cc_raw[:7]

            is_prof = nd_row[:2] == "31" and nd_row[:8] not in ND_EXCL_PROF
            sup_fdb = _is_superavit_fundeb(r, co_row, fr_suf)

            def _acc(bucket):
                bucket[sub_key]["emp"] += saldo
                if cc7 in ("6221303", "6221304", "6221307"):
                    bucket[sub_key]["liq"] += saldo
                if cc7 == "6221304":
                    bucket[sub_key]["pago"] += saldo

            def _acc_ind(b):
                b["emp"] += saldo
                if cc7 in ("6221303", "6221304", "6221307"):
                    b["liq"] += saldo
                if cc7 == "6221304":
                    b["pago"] += saldo

            # ── Quadro 10 ──────────────────────────────────────────────────
            # FR fundeb tem prioridade: CO=1001 com FR=540 ainda é Quadro 10.
            # sup_fdb: FUNDEB-superávit da UO 18903 rotulado 1001 pelo CASE.
            if fr_suf in FR_FUNDEB or co_row == "1070" or sup_fdb:
                grupo = "prof" if is_prof else "outras"
                _acc(desp_real_1070[mes][grupo])
                # L19(u)/(v): parcela do FUNDEB custeada com superávit (cofonte
                # de exercícios anteriores, prefixo '3') = superávit aplicado.
                if str(r.get("cofonte") or "")[:1] == "3":
                    superavit_aplicado[mes] += saldo
            elif co_row == "1001":
                _acc(desp_real_1001[mes])

            # ── Indicadores (linhas 11-14) ─────────────────────────────────
            # Acumula despesas de execução por tipo de fonte FUNDEB
            nd_ok = nd_row[:8] not in ND_EXCL_PROF  # exclui 31900100 e 31900300

            if fr_suf == "540" and nd_ok:
                _acc_ind(ind_real[mes]["l11_540"])
            elif fr_suf in ("541", "546") and nd_ok:
                _acc_ind(ind_real[mes]["l11_541"])
            elif fr_suf == "542" and nd_ok:
                _acc_ind(ind_real[mes]["l11_542"])
                if subfunc == "365":           # linha 13: VAAT educação infantil
                    _acc_ind(ind_real[mes]["l13"])
                if nd_row[:1] == "4":          # linha 14: VAAT capital
                    _acc_ind(ind_real[mes]["l14"])
            elif fr_suf == "543" and nd_ok:
                _acc_ind(ind_real[mes]["l11_543"])

            # linha 12: profissionais FR 1540+1541+1542 exatos; ND 31 exceto exclusões
            # CO não é filtrado aqui pois no banco essas linhas têm CO=1001
            # (incategoria nula não dispara o CASE CO=1070 no SQL)
            if fr_raw in ("1540","1541","1542") and is_prof:
                _acc_ind(ind_real[mes]["l12"])

            # ── Quadro 20 ─────────────────────────────────────────────────
            # CO=1001, FR 500/502/718, nd_ok
            #   cofuncao=12 → subfunções individuais + catchall "outras"
            #   cofuncao=28 + cosubfuncao 843/844 → "outras" (mesmo que 10.2.8)
            # Fatias com ND em ND8_EXCL_MDE (desdobramento excluído pelo RREO
            # oficial) ficam FORA de emp/liq/pago e têm o empenhado ABATIDO da
            # dotação — sem exigir FR_Q20, pois o oficial filtra por fonte
            # resumida (já garantida pelo CO=1001) e não por COFONTEFEDERAL.
            cofunc = str(r.get("cofuncao") or "").strip()
            if co_row == "1001" and nd_ok and not sup_fdb and \
               (fr_suf in FR_Q20 or nd_row[:8] in ND8_EXCL_MDE):
                if cofunc == "12":
                    q20k = "transp" if subfunc == Q20_TRANSP \
                           else (subfunc if subfunc in Q20_SUBS else "outras")
                elif cofunc == "28" and subfunc in ("843","844"):
                    q20k = "outras"
                else:
                    q20k = None
                if q20k:
                    b = desp_real_q20[mes][q20k]
                    if nd_row[:8] in ND8_EXCL_MDE:
                        # todo saldo em 62213 é movimento de empenho do mês
                        b["dot_atu"] -= saldo
                        if fr_suf not in FR_Q20:
                            # fatia de fonte fora do universo FR (ex.: 1021/
                            # 1002): nunca entrou em emp/liq/pago, mas o RREO
                            # oficial a ABATE dessas colunas — subtrair
                            # (fatias dentro do universo FR basta não somar)
                            b["emp"] -= saldo
                            if cc7 in ("6221303","6221304","6221307"):
                                b["liq"] -= saldo
                            if cc7 == "6221304":
                                b["pago"] -= saldo
                    else:
                        b["emp"] += saldo
                        if cc7 in ("6221303","6221304","6221307"):
                            b["liq"] += saldo
                        if cc7 == "6221304":
                            b["pago"] += saldo

            # ── Quadro 21 ─────────────────────────────────────────────────
            # Mesmo tratamento de ND8_EXCL_MDE do Quadro 20: fatia excluída
            # fica fora de emp/liq/pago e tem o empenhado abatido da dotação.
            # Diferença: aqui as fatias de fontes 1021/1002 (FR 1540) JÁ estão
            # na base via FR_FUNDEB — para essas basta não somar; a subtração
            # só se aplica a fatia fora da base do quadro.
            q21_base = (co_row == "1001" and fr_suf in FR_Q20) or fr_suf in FR_FUNDEB \
                       or sup_fdb
            q21_nd8  = co_row == "1001" and nd_row[:8] in ND8_EXCL_MDE
            if cofunc == "12" and nd_ok and (q21_base or q21_nd8):
                coproj_q21 = str(r.get("coprojeto")   or "").strip()
                cosubt_q21 = str(r.get("cosubtitulo") or "").strip()
                q21k = _q21_key(subfunc, coproj_q21, cosubt_q21)
                if q21k:
                    b21 = desp_real_q21[mes][q21k]
                    if q21_nd8:
                        # todo saldo em 62213 é movimento de empenho do mês
                        b21["dot_atu"] -= saldo
                        if not q21_base:
                            b21["emp"] -= saldo
                            if cc7 in ("6221303", "6221304", "6221307"):
                                b21["liq"] -= saldo
                            if cc7 == "6221304":
                                b21["pago"] -= saldo
                    else:
                        b21["emp"] += saldo
                        if cc7 in ("6221303", "6221304", "6221307"):
                            b21["liq"] += saldo
                        if cc7 == "6221304":
                            b21["pago"] += saldo

        elif (cc_raw in CC9_AC or cc_raw in CC9_AE
              or cc_raw in CC9_AD_EXTRA or cc_raw in CC9_AF):
            # ── Quadro 30: Restos a Pagar de despesas com MDE ─────────────────
            co_row = str(r.get("co") or "").strip()
            qk = _q30_grp(co_row, fr_suf)
            if qk:
                if cc_raw in CC9_AC:
                    # saldo inicial: inmes=0 — entra no mês 1
                    mes30 = mes if 1 <= mes <= 12 else 1
                    q30_acc[mes30][qk]["ac"] += saldo
                elif 1 <= mes <= 12:
                    # ad/ae/af são movimento do período (inmes 1-12) — o
                    # saldo de abertura (inmes=0) dessas contas espelha o
                    # (ac) e não deve entrar aqui (senão duplica o ac).
                    b30 = q30_acc[mes][qk]
                    if cc_raw in CC9_AE:
                        b30["ae"] += saldo
                    if cc_raw in CC9_AD_EXTRA:
                        b30["ad_extra"] += saldo
                    if cc_raw in CC9_AF:
                        b30["af"] += saldo

    if max_mes == 0:
        log.warning("  Mínimo Educação: nenhum dado de receita realizada encontrado.")
        return {"ano": ano_atual, "max_mes": 0, "previsao": {}, "receita": {}, "fundeb": {}}

    receita = {
        str(m): {k: round(rec_real[m][k], 2) for k in REC_KEYS}
        for m in range(1, max_mes + 1)
    }
    fundeb = {
        str(m): {k: round(fund_real[m][k], 2) for k in FUNDEB_KEYS}
        for m in range(1, max_mes + 1)
    }

    # ── Despesa: montar dicts de saída ───────────────────────────────────────
    # dot_atu agora está dentro de despesa[mes] (mesma dimensão que emp/liq/pago)
    def _round_sub(bucket):
        return {s: {k: round(v, 2) for k, v in vals.items()}
                for s, vals in bucket.items()}

    despesa = {
        str(m): {
            "1001": _round_sub(desp_real_1001[m]),
            "1070": {g: _round_sub(desp_real_1070[m][g]) for g in GRUPOS_1070},
        }
        for m in range(1, max_mes + 1)
    }

    q20 = {
        str(m): _round_sub(desp_real_q20[m])
        for m in range(1, max_mes + 1)
    }

    q21 = {
        str(m): _round_sub(desp_real_q21[m])
        for m in range(1, max_mes + 1)
    }

    # Quadro 30: por mês, como os demais quadros. O ag (saldo final =
    # ac - ae - af) é calculado no dashboard sobre o período agregado.
    # ad = ae + ad_extra — ver comentário de CC9_AD_EXTRA acima.
    def _q30_row(b):
        ad = b["ae"] + b["ad_extra"]
        return {"ac": round(b["ac"], 2), "ad": round(ad, 2),
                "ae": round(b["ae"], 2), "af": round(b["af"], 2)}

    q30 = {
        str(m): {k: _q30_row(q30_acc[m][k]) for k in Q30_KEYS}
        for m in range(1, max_mes + 1)
    }

    indicadores = {
        str(m): {k: {f: round(v, 2) for f, v in vals.items()}
                 for k, vals in ind_real[m].items()}
        for m in range(1, max_mes + 1)
    }

    log.info(f"  Mínimo Educação: max_mes={max_mes}")
    return {
        "ano":             ano_atual,
        "max_mes":         max_mes,
        "previsao":        {k: round(rec_prev[k], 2) for k in REC_KEYS},
        "fundeb_previsao": {k: round(fund_prev[k], 2) for k in FUNDEB_KEYS},
        "receita":         receita,
        "fundeb":          fundeb,
        "despesa":         despesa,
        "q20":             q20,
        "q21":             q21,
        "q30":             q30,
        "indicadores":     indicadores,
        # L19 — Aplicação do Superávit de Exercício Anterior (Art.25 §3 LC 14.113).
        # (s)/(t) são constantes de apuração; (u)/(v)/(w)/(x) e L25 são derivados
        # no dashboard a partir de superavit_aplicado (por período selecionado):
        #   (u) = Σ aplicado[m], m≤4 ; (v) = Σ aplicado[m], m>4
        #   (w) = max(0, t−u−v) ; (x) = max(0, t−u) ; L25 = (x)
        "superavit": (lambda s191, s192: {
            # L19(s) = 10% × receita FUNDEB ano anterior (19.1/19.2)
            "ant":      round(s191, 2),
            "s_192":    round(s192, 2),
            # L8.1 por bimestre (re-apuração) + L8.2. O dashboard escolhe o L8.1
            # vigente conforme o bimestre selecionado; (t) = L8.1(bim) + L8.2.
            "l8_1_por_bim": _l81_por_bim(),
            "l8_2":     round(MDE_SUPERAVIT_8_2, 2),
            "residual": round(MDE_SUPERAVIT_8_2, 2),
            "t_192": 0.0,
        })(fundeb_ant["f61"] * 0.10, fundeb_ant["f_1715"] * 0.10),
        # Superávit do FUNDEB aplicado por mês (empenho cofonte 3xx) → (u)/(v)
        "superavit_aplicado": {
            str(m): round(superavit_aplicado[m], 2) for m in range(1, max_mes + 1)
        },
    }


def save_minimo_educacao_gz(D_obj):
    payload = {"atualizado_em": datetime.now(timezone.utc).isoformat()}
    payload.update(D_obj)
    content = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    gz_path = GZ_DIR / "minimo_educacao.json.gz"
    with gzip.open(gz_path, "wb", compresslevel=9) as f:
        f.write(content)
    size_kb = gz_path.stat().st_size / 1024
    log.info(f"  minimo_educacao.json.gz -- {size_kb:.1f} KB")


def upsert_minimo_educacao_supabase(D_obj):
    if not SUPABASE_URL or not SUPABASE_KEY:
        log.warning("  Supabase: nao configurado. Pulando upsert minimo_educacao.")
        return
    try:
        ano = D_obj.get("ano") or datetime.now().year
        payload = [{
            "ano":           ano,
            "dados":         D_obj,
            "atualizado_em": datetime.now(timezone.utc).isoformat(),
        }]
        total = _supabase_upsert("minimo_educacao", payload, "ano")
        log.info(f"  Supabase: minimo_educacao {ano} enviada ({total} linha).")
    except Exception as e:
        log.error(f"  Supabase minimo_educacao falhou: {type(e).__name__}: {e}")


def init_oracle():
    import oracledb
    if CLIENT_PATH:
        log.info(f"Inicializando thick mode -> {CLIENT_PATH}")
        oracledb.init_oracle_client(lib_dir=CLIENT_PATH)
    else:
        log.info("Usando thin mode (sem Oracle Client local)")
    return oracledb


def run():
    try:
        oracledb = init_oracle()
    except ImportError:
        raise ImportError("Execute: pip install oracledb")

    if not DB_USER or not DB_PASSWORD:
        raise ValueError("DB_USER e DB_PASSWORD precisam estar definidos no .env")

    log.info(f"Conectando ao Oracle -> {DB_DSN}  [schema: {SCHEMA_ANO}]")

    pool = oracledb.create_pool(
        user=DB_USER, password=DB_PASSWORD, dsn=DB_DSN,
        min=DB_MIN, max=DB_MAX, increment=DB_INC,
    )

    with pool.acquire() as conn:
        log.info("Conexao estabelecida. Iniciando extracao...")
        with conn.cursor() as cur:
            for item in QUERIES:
                log.info(f"Extraindo -> {item['file']}")
                try:
                    data = fetch(cur, resolve_query(item))
                    if item.get("transform") == "rcl":
                        D_obj = build_rcl_data(data)
                        save_rcl_gz(D_obj)
                        upsert_rcl_supabase(D_obj)
                        save_json(item["file"], data)
                    elif item.get("transform") == "restos_a_pagar":
                        registros = build_restos_a_pagar_data(data)
                        save_restos_a_pagar_gz(registros)
                        upsert_restos_a_pagar_supabase(registros)
                        save_json(item["file"], data)
                    elif item.get("transform") == "resultado_primario_nominal":
                        D_obj = build_resultado_primario_nominal_data(data)
                        save_resultado_primario_nominal_gz(D_obj)
                        upsert_resultado_primario_nominal_supabase(D_obj)
                        save_json(item["file"], data)
                    elif item.get("transform") == "poupanca_corrente":
                        D_obj = build_poupanca_corrente_data(data)
                        save_poupanca_corrente_gz(D_obj)
                        upsert_poupanca_corrente_supabase(D_obj)
                        save_json(item["file"], data)
                    elif item.get("transform") == "minimo_saude":
                        D_obj = build_minimo_saude_data(data)
                        save_minimo_saude_gz(D_obj)
                        upsert_minimo_saude_supabase(D_obj)
                        save_json(item["file"], data)
                    elif item.get("transform") == "minimo_educacao":
                        D_obj = build_minimo_educacao_data(data)
                        save_minimo_educacao_gz(D_obj)
                        upsert_minimo_educacao_supabase(D_obj)
                        save_json(item["file"], data)
                    elif item["file"] == "receita.json":
                        save_json(item["file"], data)
                        save_json_gz(item["file"], data)
                        upsert_receita_supabase(data)
                    elif item["file"] == "despesa.json":
                        save_json(item["file"], data)
                        save_json_gz(item["file"], data)
                        upsert_despesa_supabase(data)
                    else:
                        save_json(item["file"], data)
                        save_json_gz(item["file"], data)
                except Exception as e:
                    log.error(f"  Erro em {item['file']}: {type(e).__name__}: {e}")
                    import traceback
                    traceback.print_exc()

    pool.close()
    log.info("ETL concluido com sucesso.")


if __name__ == "__main__":
    run()
